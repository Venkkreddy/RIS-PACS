import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Define Paths
BASE_DIR = r"D:\TDAI\ris-pacs"
JSON_PATH = os.path.join(BASE_DIR, "trivitrion_workstation_benchmark_v4.json")
EXCEL_OUTPUT = os.path.join(BASE_DIR, "TDAI_RIS_PACS_PRD_Roadmap_Suite.xlsx")
PDF_OUTPUT = os.path.join(BASE_DIR, "TDAI_RIS_PACS_Product_Roadmap_PRD.pdf")

# Brand Colors (Hex and RGB)
TEAL_HEX = "00B4A6"
NAVY_HEX = "1A2B56"
GRAY_HEX = "F3F4F6"
WHITE_HEX = "FFFFFF"

TEAL_RGB = (0, 180, 166)
NAVY_RGB = (26, 43, 86)
GRAY_RGB = (243, 244, 246)
DARK_GRAY_RGB = (80, 80, 80)
BLACK_RGB = (17, 24, 39)

def generate_excel():
    print("Generating Excel sheet...")
    # Load JSON Benchmark Data
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Common Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1A2B56")
    section_font = Font(name=font_family, size=12, bold=True, color="00B4A6")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    bold_data_font = Font(name=font_family, size=10, bold=True)
    italic_data_font = Font(name=font_family, size=9, italic=True)

    header_fill = PatternFill(start_color=TEAL_HEX, end_color=TEAL_HEX, fill_type="solid")
    navy_fill = PatternFill(start_color=NAVY_HEX, end_color=NAVY_HEX, fill_type="solid")
    zebra_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")
    
    thin_border_side = Side(style='thin', color='D1D5DB')
    double_border_side = Side(style='double', color='1A2B56')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=thin_border_side, bottom=double_border_side)

    # 1. Summary Dashboard Sheet
    ws_dash = wb.create_sheet("Summary Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.cell(row=1, column=1, value="TDAI RAD Platform Workstation Feature Benchmark").font = title_font
    ws_dash.cell(row=2, column=1, value="Executive Summary & Core Platform Overview").font = italic_data_font
    
    # KPI Grid
    kpi_headers = ["Key Metric", "Value", "Status Details", "Roadmap Focus"]
    for col_idx, text in enumerate(kpi_headers, 1):
        cell = ws_dash.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_dash.row_dimensions[4].height = 25

    kpi_data = [
        ["Total Features Evaluated", "68", "Comprehensive audit covering 9 workflow modules", "Benchmark Baseline"],
        ["Must-Have Features", "37", "Core radiological capabilities & requirements", "Immediate Verification"],
        ["Good-to-Have Features", "19", "Workflow optimization & productivity tools", "Secondary Enhancements"],
        ["Differentiating Features", "12", "Strategic AI models & clinical assistant tools", "Primary Competitive Edge"],
        ["Fully Implemented / Working", "7", "Production-ready clinical features", "Validate & Maintain"],
        ["Supported (Basic / Standard Tools)", "48", "Features running basic logic or standard packages", "UX/UI Refining"],
        ["Planned (Not Supported / Roadmap Target)", "13", "High-differentiating AI models & offline tasks", "Strategic R&D Investments"]
    ]

    for row_idx, row_data in enumerate(kpi_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_dash.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
            if col_idx == 2:
                cell.font = bold_data_font
                cell.alignment = Alignment(horizontal="center")
        ws_dash.row_dimensions[row_idx].height = 20

    # Project Objectives Box
    ws_dash.cell(row=14, column=1, value="Platform Objectives").font = section_font
    obj_data = [
        ["Product Vision", "Provide a unified, cloud-native RIS/PACS monorepo workstation featuring seamless AI-assisted imaging workflow and local medical dictation, deployable on-premise or cloud in under 3 minutes."],
        ["Core Goals", "1. Zero Typing Errors: Connect check-in scheduling records straight to scans using MWL protocol.\n2. TAT Reduction: 40% improvement in reporting using embedded viewer and voice-to-report pipelines.\n3. Zero-Config Launch: Electron setup packages SQL database, PACS stack, AI modules, and viewer in a single installer."]
    ]
    for r_idx, (k, v) in enumerate(obj_data, 15):
        cell_k = ws_dash.cell(row=r_idx, column=1, value=k)
        cell_k.font = bold_data_font
        cell_k.border = thin_border
        cell_k.fill = zebra_fill
        
        cell_v = ws_dash.cell(row=r_idx, column=2, value=v)
        cell_v.font = data_font
        cell_v.border = thin_border
        cell_v.alignment = Alignment(wrap_text=True, vertical="top")
        ws_dash.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=4)
        ws_dash.row_dimensions[r_idx].height = 50

    # 2. Market Share & Competitors Sheet
    ws_mkt = wb.create_sheet("Market & Competitors")
    ws_mkt.views.sheetView[0].showGridLines = True
    ws_mkt.cell(row=1, column=1, value="RIS/PACS Market Share & Competitor Analysis").font = title_font
    ws_mkt.cell(row=2, column=1, value="Target Market: Standalone Diagnostic Centers & Mid-Tier Regional Hospitals").font = italic_data_font

    mkt_headers = ["Competitor Group", "Est. Market Share (%)", "Target Segments", "Strengths", "Weaknesses / Gaps vs TDAI"]
    for col_idx, text in enumerate(mkt_headers, 1):
        cell = ws_mkt.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_mkt.row_dimensions[4].height = 25

    mkt_data = [
        ["Global MNCs (GE Centricity, Siemens, Philips)", "45.0%", "Enterprise Hospitals", "Robust FDA clearance, brand equity, multi-department networks", "Extremely expensive, complex on-site setups (days to configure), slow custom integrations"],
        ["MedSynaptics", "20.0%", "Mid-Tier Clinics & Regional Hospitals", "Established regional PACS client base, modular licensing", "Lacks embedded local voice-to-text dictation, AI tools are add-on cloud services (costly latency)"],
        ["FilmPlus (Telesofia / Local Vendors)", "12.0%", "Local Standalone Diagnostics", "Low cost, solid printing/film layouts integration, long presence", "Archaic UI, lacks advanced web viewer (non-zero footprint), zero integrated AI models"],
        ["Other Local / In-house Custom", "21.0%", "Small Diagnostics & Labs", "Extremely cheap, highly customized to individual doctors", "No security compliance (HIPAA), poor support, high crash rates, no scalability"],
        ["TDAI (Trivitron Healthcare Target)", "2.0% (Current)", "Standalone Diagnostics & Mid-Tier Chains", "Monorepo React stack, built-in local MONAI and MedASR servers, 3-minute setup Electron desktop package", "New market entrant, brand validation in progress (target 8% share in 18 months)"]
    ]

    for row_idx, row_data in enumerate(mkt_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_mkt.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
            if col_idx == 2:
                cell.font = bold_data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_mkt.row_dimensions[row_idx].height = 40

    # 3. Product Roadmap Sheet
    ws_rd = wb.create_sheet("Product Roadmap")
    ws_rd.views.sheetView[0].showGridLines = True
    ws_rd.cell(row=1, column=1, value="TDAI Workstation Releases & Roadmap").font = title_font
    ws_rd.cell(row=2, column=1, value="Strategic Release Schedule (Q3 2026 - Q2 2027)").font = italic_data_font

    rd_headers = ["Phase", "Release Target", "Core Objective", "Key Deliverables", "Status", "Target Audience"]
    for col_idx, text in enumerate(rd_headers, 1):
        cell = ws_rd.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_rd.row_dimensions[4].height = 25

    rd_data = [
        ["Phase 1: Foundation", "Q3 2026", "Core RIS/PACS & monorepo stability", "React frontend, Express backend, Orthanc PACS server integrations, single-click Electron desktop packaging, basic HIPAA auditing.", "Completed / Local QA", "Radiographers & Receptionists"],
        ["Phase 2: Workflow Parity", "Q4 2026", "Streamline user interfaces & configurations", "Multi-column sorting, customizable worklist columns, study metadata preview side-panel, auto-retry transfer queue for C-STORE.", "In Planning", "Clinical Technicians"],
        ["Phase 3: AI & Smart Reports", "Q1 2027", "Introduce AI diagnostic & voice reporting features", "Integrated local voice-to-text dictation (MedASR) with LLM correction, MONAI AI model priors, auto critical-findings worklist priority.", "Under Active Dev", "Radiologists"],
        ["Phase 4: Enterprise Scale", "Q2 2027", "Hospital groups and printing automation", "DICOM print layout templates UI, custom mouse/keyboard shortcut mapping, multi-PACS query load balancing, auto spine/leg image stitching.", "Planned", "Hospital IT & Lead Radiologists"]
    ]

    for row_idx, row_data in enumerate(rd_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_rd.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
            if col_idx in [1, 2, 5]:
                cell.font = bold_data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_rd.row_dimensions[row_idx].height = 50

    # 4. Gapfit Analysis Sheet
    ws_gap = wb.create_sheet("Gapfit Summary")
    ws_gap.views.sheetView[0].showGridLines = True
    ws_gap.cell(row=1, column=1, value="Competitor Feature Gapfit Analysis").font = title_font
    ws_gap.cell(row=2, column=1, value="Strategic Feature Gaps, Priority, & Roadmap Remediation").font = italic_data_font

    gap_headers = ["Identified Gap Feature", "Module", "Competitor Capability", "TDAI Current Status", "Business Priority", "Roadmap Remediation"]
    for col_idx, text in enumerate(gap_headers, 1):
        cell = ws_gap.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_gap.row_dimensions[4].height = 25

    gap_data = [
        ["Customizable Columns", "Patient Worklist", "MedSynaptics: User can toggle 19+ fields.", "Supported (Fixed Grid) only.", "High", "Phase 2 (Q4 2026): Add drag-and-drop column configurator UI in settings."],
        ["Multi-column Sorting", "Patient Worklist", "MedSynaptics: Interactive sorting across multiple headers.", "Supported (Priority-first automatic only).", "High", "Phase 2 (Q4 2026): Implement multi-key table state sorting in TanStack Table."],
        ["Interactive Print Layout", "Printing", "FilmPlus: Choose grid layout, film density, and custom frame grids.", "Supported via Weasis app launcher. Web client only prints PDF reports.", "High", "Phase 4 (Q2 2027): Integrate print-template canvas manager directly in OHIF viewer wrapper."],
        ["Failed DICOM Queue", "DICOM Networking", "FilmPlus: User interface to monitor, pause, and retry failed sends.", "Background automatic retry only; no user interface panel.", "Medium", "Phase 2 (Q4 2026): Build a detailed DICOM Transfer Monitor tab in Developer Portal."],
        ["Custom Shortcut Mapping", "Settings & Personalization", "MedSynaptics: Fully maps user mouse click and wheel tools.", "Fixed preset keyboard shortcuts and mouse clicks in OHIF config.", "Medium", "Phase 3 (Q1 2027): Create hotkey profile configuration interface in Settings Page."],
        ["Spine / Leg Image Stitching", "Viewer Tools", "FilmPlus: Auto-stitch overlapping X-rays into a single image.", "Not supported in web viewer client.", "High", "Phase 4 (Q2 2027): Build canvas-based image stitching module in OHIF web client."]
    ]

    for row_idx, row_data in enumerate(gap_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_gap.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
            if col_idx in [5, 4]:
                cell.font = bold_data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_gap.row_dimensions[row_idx].height = 45

    # 5. Populate Feature Benchmarking Sheets from JSON (all 68 features)
    feature_sheets = [
        "Patient Worklist", "Acquisition & Study Mgmt", "Viewer Tools", 
        "Measurements", "Annotation Tools", "DICOM Networking", 
        "Printing", "Settings & Personalization", "AI & Smart Features"
    ]
    
    for sh_name in feature_sheets:
        if sh_name not in data:
            continue
        
        ws = wb.create_sheet(sh_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Header Rows
        ws.cell(row=1, column=1, value=f"Module: {sh_name}").font = title_font
        ws.cell(row=2, column=1, value=data[sh_name][1][0] if len(data[sh_name]) > 1 else "").font = italic_data_font
        
        # Grid header at row 4
        headers = ["Feature", "Classification", "Description & Detailing", "FilmPlus", "MedSynaptics", "Current Trivitron (RAD) State", "Recommendation / Enhancement", "Priority", "Business / Clinical Benefit"]
        for col_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[4].height = 25
        
        # Rows start at 5
        curr_row = 5
        # Skip the first 4 elements which were module title & headers
        for r_data in data[sh_name][4:]:
            if len(r_data) < 2 or not r_data[0]:
                continue
            
            for col_idx, val in enumerate(r_data, 1):
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Check status and color/style appropriately
                if col_idx in [1, 2, 8]:
                    cell.font = bold_data_font
                
                # Zebra colors
                if curr_row % 2 == 0:
                    cell.fill = zebra_fill
                    
            ws.row_dimensions[curr_row].height = 65
            curr_row += 1

    # Auto-adjust column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            if max_len > 45:
                ws.column_dimensions[col_letter].width = 35
            elif max_len < 10:
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(EXCEL_OUTPUT)
    print(f"Excel file successfully generated at {EXCEL_OUTPUT}")


class PRD_PDF_Class(FPDF):
    def header(self):
        if self.page_no() > 1:
            self.set_fill_color(TEAL_RGB[0], TEAL_RGB[1], TEAL_RGB[2])
            self.rect(0, 0, 210, 15, "F")
            
            self.set_text_color(255, 255, 255)
            self.set_font("Helvetica", "B", 9)
            self.cell(0, -5, "TDAI RIS/PACS Smart Radiology Workstation - PRD & Roadmap", align="R")
            
            self.ln(12)
            
    def footer(self):
        if self.page_no() > 1:
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(DARK_GRAY_RGB[0], DARK_GRAY_RGB[1], DARK_GRAY_RGB[2])
            self.cell(0, 10, "Trivitron Healthcare & TDAI Digital Solutions -- Confidential", align="L")
            self.set_x(-40)
            self.cell(0, 10, f"Page {self.page_no()}", align="R")

def add_chapter_title(pdf, text):
    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(NAVY_RGB[0], NAVY_RGB[1], NAVY_RGB[2])
    pdf.cell(0, 8, text, ln=True)
    x = pdf.get_x()
    y = pdf.get_y()
    pdf.set_draw_color(TEAL_RGB[0], TEAL_RGB[1], TEAL_RGB[2])
    pdf.set_line_width(0.8)
    pdf.line(x, y, x + 190, y)
    pdf.ln(5)

def add_chapter_body(pdf, text):
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(BLACK_RGB[0], BLACK_RGB[1], BLACK_RGB[2])
    pdf.multi_cell(0, 5, text)
    pdf.ln(3)

def add_bullet(pdf, title, desc):
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(TEAL_RGB[0], TEAL_RGB[1], TEAL_RGB[2])
    pdf.write(5, chr(149) + " " + title + ": ")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(BLACK_RGB[0], BLACK_RGB[1], BLACK_RGB[2])
    pdf.write(5, desc + "\n")
    pdf.ln(1)

def add_callout(pdf, title, body, style="info"):
    pdf.set_font("Helvetica", "B", 10)
    if style == "warning":
        fill = (254, 243, 199)
        border = (217, 119, 6)
        text_color = (180, 83, 9)
    else:
        fill = (224, 242, 254)
        border = (14, 116, 144)
        text_color = (21, 94, 117)
        
    x = pdf.get_x()
    y = pdf.get_y()
    
    pdf.set_fill_color(fill[0], fill[1], fill[2])
    pdf.set_draw_color(border[0], border[1], border[2])
    pdf.set_line_width(0.5)
    pdf.set_text_color(text_color[0], text_color[1], text_color[2])
    
    lines = len(body.split('\n')) + 1
    box_height = max(12, lines * 4.5 + 6)
    
    pdf.rect(x, y, 190, box_height, "DF")
    pdf.set_xy(x + 4, y + 2)
    pdf.cell(0, 4, title, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_xy(x + 4, y + 6)
    pdf.multi_cell(182, 4.2, body)
    
    pdf.set_xy(x, y + box_height + 4)

def add_table(pdf, headers, col_widths, rows_data):
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(TEAL_RGB[0], TEAL_RGB[1], TEAL_RGB[2])
    pdf.set_text_color(255, 255, 255)
    pdf.set_draw_color(209, 213, 223)
    pdf.set_line_width(0.3)
    
    for header, width in zip(headers, col_widths):
        pdf.cell(width, 7, header, border=1, align="C", fill=True)
    pdf.ln()
    
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(BLACK_RGB[0], BLACK_RGB[1], BLACK_RGB[2])
    
    row_count = 0
    for row in rows_data:
        row_count += 1
        if row_count % 2 == 0:
            pdf.set_fill_color(GRAY_RGB[0], GRAY_RGB[1], GRAY_RGB[2])
        else:
            pdf.set_fill_color(255, 255, 255)
            
        for cell_idx, (cell_val, width) in enumerate(zip(row, col_widths)):
            is_last = (cell_idx == len(row) - 1)
            pdf.cell(width, 7, str(cell_val), border=1, fill=True, ln=is_last)
            
    pdf.ln(3)

def generate_pdf():
    print("Generating PDF PRD document...")
    pdf = PRD_PDF_Class()
    pdf.set_margins(10, 15, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # ------------------ COVER PAGE ------------------
    pdf.set_fill_color(NAVY_RGB[0], NAVY_RGB[1], NAVY_RGB[2])
    pdf.rect(0, 0, 210, 297, "F")
    
    pdf.set_fill_color(TEAL_RGB[0], TEAL_RGB[1], TEAL_RGB[2])
    pdf.rect(0, 100, 210, 20, "F")
    
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_xy(15, 60)
    pdf.cell(0, 10, "TDAI RIS / PACS WORKSTATION", ln=True)
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(TEAL_RGB[0], TEAL_RGB[1], TEAL_RGB[2])
    pdf.set_xy(15, 72)
    pdf.cell(0, 10, "SMART RADIOLOGY WORKSTATION SUITE", ln=True)
    
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_xy(15, 106)
    pdf.cell(0, 10, "Product Requirements Document (PRD) & Strategic Roadmap", ln=True)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.set_xy(15, 240)
    pdf.cell(0, 5, "Author: TDAI Product Engineering Team", ln=True)
    pdf.cell(0, 5, "Company: Trivitron Healthcare & TDAI Digital Solutions", ln=True)
    pdf.cell(0, 5, "Date: July 14, 2026", ln=True)
    pdf.cell(0, 5, "Version: 1.0.0 (Release-Ready Spec)", ln=True)
    pdf.cell(0, 5, "Classification: Commercial Confidential", ln=True)
    
    # ------------------ CHAPTER 1 ------------------
    pdf.add_page()
    add_chapter_title(pdf, "1. Executive Summary & Core Objectives")
    
    summary_text = (
        "Radiology diagnostic facilities worldwide experience severe bottlenecks and clinical mismatches due to disjointed software platforms. "
        "The Receptionist database, the technician modality console scanner, and the radiologist report drafting systems operate as separate silos. "
        "Technicians spend hours retyping patient names, ages, and medical record numbers (MRNs) directly onto CT/X-Ray consoles, resulting in "
        "transcription errors, patient study mismatches, billing discrepancies, and medical compliance breaches.\n\n"
        "TDAI RIS/PACS solves these bottlenecks by delivering a unified, multi-tenant workstation suite. It packages a React-based clinical web panel, "
        "a custom-configured Orthanc PACS server supporting DICOM Modality Worklist (MWL) and Modality Performed Procedure Step (MPPS) protocols, "
        "a Lucene index search engine (Dicoogle), a HIPAA-compliant SHA-256 hash-chained audit database, and a high-fidelity diagnostic web viewer (OHIF) "
        "into a single-click Electron desktop application.\n\n"
        "This integration achieves immediate business benefits:"
    )
    add_chapter_body(pdf, summary_text)
    
    add_bullet(pdf, "Zero Demographics Mismatches", "Check-ins synchronise instantly from the RIS database to the scanner via DICOM MWL query.")
    add_bullet(pdf, "Real-time Progress Tracker", "Scanner exposures automatically notify status changes back to the clinical dashboard via MPPS.")
    add_bullet(pdf, "Accelerated Report Turnaround Time", "Embedded diagnostic OHIF viewer runs side-by-side with an AI-powered voice-dictation transcription engine.")
    
    add_callout(pdf, "Platform Success Criteria & Target Metrics", 
                "1. Zero Typing Mismatches: 100% data entry consistency via MWL query on scanners.\n"
                "2. Rapid Installation: Installer installs and starts the full 8-container local Docker stack in under 3 minutes.\n"
                "3. TAT Improvement: At least a 40% reduction in report drafting time using integrated speech dictation.",
                "info")
                
    # ------------------ CHAPTER 2 ------------------
    add_chapter_title(pdf, "2. Market Share & Competitor Benchmarking")
    
    mkt_intro = (
        "The mid-market diagnostic sector in India and emerging regions is currently dominated by localized legacy clients. "
        "While enterprise hospitals deploy expensive, custom PACS networks from major MNCs, standalone diagnostic clinics and mid-tier regional hospitals "
        "require a lightweight, affordable, yet high-performance clinical imaging platform. "
        "TDAI's strategic market benchmarking compares the platform with two major regional competitors: MedSynaptics and FilmPlus."
    )
    add_chapter_body(pdf, mkt_intro)
    
    headers = ["Competitor Group", "Market Share", "Target Segment", "Key Advantage"]
    widths = [50, 25, 45, 70]
    mkt_rows = [
        ["Global MNCs", "45.0%", "Enterprise Hospitals", "Robust FDA approvals, broad contract reach"],
        ["MedSynaptics", "20.0%", "Mid-Tier Clinics & PACS", "Established local client footprint, modular"],
        ["FilmPlus (Telesofia/Local)", "12.0%", "Small Diagnostic Centers", "Low cost, long-standing film prints"],
        ["Other Local / In-house", "21.0%", "Independent Labs", "Archaic custom viewer, minimal upfront cost"],
        ["TDAI (Trivitron Target)", "2.0% (Current)", "Standalone Diagnostics & Chains", "Built-in local AI, unified web workstation"]
    ]
    add_table(pdf, headers, widths, mkt_rows)
    
    # ------------------ CHAPTER 3 ------------------
    pdf.add_page()
    add_chapter_title(pdf, "3. Gapfit Analysis & Feature Audit")
    
    gap_intro = (
        "A rigorous audit comparing TDAI RIS/PACS against competitors across 68 features in 9 core workspaces revealed a "
        "highly competitive baseline. TDAI supports 80%+ of standard radiological workstation features out-of-the-box via "
        "its open-source base core integrations (OHIF, Orthanc, Dicoogle, and MONAI). "
        "However, key functional gaps must be addressed to ensure competitive parity and protect our product margins."
    )
    add_chapter_body(pdf, gap_intro)
    
    add_bullet(pdf, "Customizable Columns (Priority: High)", "Competitors support 19+ togglable columns; TDAI currently uses a hardcoded data grid. Phase 2 roadmap target.")
    add_bullet(pdf, "Multi-Column Grid Sorting (Priority: High)", "Critical for sorting emergency and normal studies by date and modality simultaneously. Phase 2 target.")
    add_bullet(pdf, "Interactive Print Customization (Priority: High)", "Competitors support detailed film layouts (2x2, 3x4) and camera templates. TDAI relies on external Weasis integrations. Phase 4 target.")
    add_bullet(pdf, "Spine & Leg Image Stitching (Priority: High)", "Orthopedic scans require merging multiple series. Essential for standalone clinics. Phase 4 target.")
    add_bullet(pdf, "Failed C-STORE Transfer Queue UI (Priority: Medium)", "Visual dashboard tracking failed transfers to other hospital nodes. Phase 2 target.")
    add_bullet(pdf, "Custom Mouse/Keyboard Mapping (Priority: Medium)", "Allows radiologists to map custom viewer controls matching their muscle memory. Phase 3 target.")

    # ------------------ CHAPTER 4 ------------------
    add_chapter_title(pdf, "4. Strategic Product Roadmap & Phases")
    
    rd_intro = (
        "To systematically bridge competitor gaps and deploy high-differentiating AI models, the TDAI Product Engineering "
        "Roadmap is structured into four distinct quarterly phases from Q3 2026 to Q2 2027:"
    )
    add_chapter_body(pdf, rd_intro)
    
    rd_headers = ["Release Phase", "Target", "Core Objective", "Primary Audience"]
    rd_widths = [50, 20, 85, 35]
    rd_rows = [
        ["Phase 1: Foundation", "Q3 2026", "Express/React monorepo, Orthanc Pacs integration", "Receptionist / Tech"],
        ["Phase 2: Workflow Parity", "Q4 2026", "Grid column selection, multi-sorting, preview pane", "Technician / Lead Tech"],
        ["Phase 3: AI & Smart Reports", "Q1 2027", "MedASR dictation, MONAI diagnostic overlays", "Radiologists"],
        ["Phase 4: Enterprise Scale", "Q2 2027", "DICOM print layout designer, multi-PACS loading", "Hospital IT / Radiologists"]
    ]
    add_table(pdf, rd_headers, rd_widths, rd_rows)

    # ------------------ CHAPTER 5 ------------------
    pdf.add_page()
    add_chapter_title(pdf, "5. Security, Audit Logs & HIPAA Compliance")
    
    compliance_text = (
        "Medical records and patient studies constitute Protected Health Information (PHI) subject to strict regulatory compliance "
        "(HIPAA in the US, DISHA in India, GDPR in Europe). TDAI workstation enforces security at the network, database, and application levels.\n\n"
        "Key security configurations integrated in the TDAI monorepo backend include:"
    )
    add_chapter_body(pdf, compliance_text)
    
    add_bullet(pdf, "SHA-256 Hash-Chained Audit Trail", "All patient creations, edits, study views, downloads, and report exports are logged to a secure, immutable log file where each entry hashes the previous log. Integrity is verified dynamically via backend API.")
    add_bullet(pdf, "Automatic Session Expiration", "Clinical sessions automatically expire and log out users after 15 minutes of inactivity to protect clinical screens from unauthorized access.")
    add_bullet(pdf, "Emergency 'Break-Glass' PHI Access", "Authorized clinicians can bypass strict patient assignments to view emergency studies, triggering high-priority audit logs sent to administrators.")
    add_bullet(pdf, "Password & Rate Limit Protections", "Enforced password strength policies, account lockout after 5 consecutive failed logins, and API request throttling limited to 300 requests/minute.")

    add_callout(pdf, "Regulatory Compliance Guidelines", 
                "WARNING: De-identifying patient names and demographics during exports is mandatory. "
                "All external clinical access points must run over secure TLS HTTPS connections. "
                "Local database archives (Postgres & Firestore logs) must be backed up daily to encrypted storage drives.",
                "warning")
                
    pdf.output(PDF_OUTPUT)
    print(f"PDF PRD document successfully generated at {PDF_OUTPUT}")

if __name__ == "__main__":
    generate_excel()
    generate_pdf()
